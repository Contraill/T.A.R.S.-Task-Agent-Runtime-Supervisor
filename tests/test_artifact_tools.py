import io
from pathlib import Path
import subprocess
import tarfile
import zipfile

import pytest

from tars import approvals, artifact_tools, document_tools, media_tools, policy, state_store


@pytest.fixture
def artifacts(monkeypatch, tmp_path):
    monkeypatch.setattr(state_store, "STATE_DB_PATH", tmp_path / "state.sqlite3")
    monkeypatch.setattr(state_store, "TASK_ROOT", tmp_path / "legacy")
    monkeypatch.setattr(state_store, "TASK_EVENTS_ROOT", tmp_path / "events")
    monkeypatch.setattr(state_store, "TASK_INDEX_PATH", tmp_path / "index")
    root = tmp_path / "workspace"
    root.mkdir()
    return root


def approve(tool, target, root):
    request = policy.ScopeRequest(tool, "write", str(target), allowed_paths=(str(root),))
    decision = policy.ScopeGuard().evaluate(request)
    broker = approvals.ApprovalBroker()
    pending = broker.request(request, decision, scope="target")
    broker.decide(pending.id, approve=True)
    return pending.id


def test_archive_round_trip_hash_and_evidence(artifacts):
    source = artifacts / "source"
    source.mkdir()
    (source / "note.txt").write_text("verified payload")
    archive = artifacts / "bundle.zip"
    tools = artifact_tools.ArchiveTools((artifacts,))
    created = tools.create(archive, (source,), approval_ids={
        "write-0": approve("archive.create", archive, artifacts),
    })
    assert created.succeeded and tools.list(archive).data["members"][0]["name"].endswith("note.txt")
    destination = artifacts / "extracted"
    extracted = tools.extract(archive, destination, approval_ids={
        "write-0": approve("archive.extract", destination, artifacts),
    })
    assert extracted.succeeded and (destination / "source" / "note.txt").read_text() == "verified payload"
    integrity = artifact_tools.IntegrityTools((artifacts,)).hash(archive, expected=created.data["sha256"])
    assert integrity.succeeded and integrity.data["verified"] and integrity.evidence_ids


@pytest.mark.parametrize("kind", ["zip", "tar"])
def test_archive_extract_rejects_traversal_before_writing_outside(artifacts, kind):
    archive = artifacts / f"bad.{kind}"
    if kind == "zip":
        with zipfile.ZipFile(archive, "w") as handle:
            handle.writestr("../escape.txt", "bad")
    else:
        with tarfile.open(archive, "w") as handle:
            value = b"bad"
            member = tarfile.TarInfo("../escape.txt")
            member.size = len(value)
            handle.addfile(member, io.BytesIO(value))
    destination = artifacts / "output"
    with pytest.raises(ValueError, match="unsafe|escapes"):
        artifact_tools.ArchiveTools((artifacts,)).extract(archive, destination, approval_ids={
            "write-0": approve("archive.extract", destination, artifacts),
        })
    assert not (artifacts.parent / "escape.txt").exists()


def test_archive_destination_symlink_swap_after_authorization_cannot_escape(
        artifacts, monkeypatch):
    archive = artifacts / "bundle.zip"
    with zipfile.ZipFile(archive, "w") as handle:
        handle.writestr("payload.txt", "contained")
    destination = artifacts / "output"
    destination.mkdir()
    outside = artifacts.parent / "outside"
    outside.mkdir()
    tools = artifact_tools.ArchiveTools((artifacts,))
    approval = approve("archive.extract", destination, artifacts)
    authorize = tools.artifacts.runtime.authorize

    def swap_after_authorization(*args, **kwargs):
        actions = authorize(*args, **kwargs)
        destination.rename(artifacts / "displaced")
        destination.symlink_to(outside, target_is_directory=True)
        return actions

    monkeypatch.setattr(tools.artifacts.runtime, "authorize", swap_after_authorization)
    with pytest.raises(OSError):
        tools.extract(archive, destination, approval_ids={"write-0": approval})
    assert not (outside / "payload.txt").exists()


def test_csv_range_write_read_and_document_edit(artifacts):
    sheet = artifacts / "data.csv"
    sheet.write_text("one,two\nthree,four\n")
    spreadsheets = document_tools.SpreadsheetTools((artifacts,))
    write = spreadsheets.write_range(sheet, "B2:B2", [["changed"]],
                                      approval_id=approve("spreadsheet.write_range", sheet, artifacts))
    assert write.succeeded
    assert spreadsheets.read_range(sheet, "A1:B2").data["values"][1] == ["three", "changed"]
    document = artifacts / "note.md"
    document.write_text("old claim")
    edited = document_tools.DocumentTools((artifacts,)).edit(
        document, [("old", "supported")],
        approval_id=approve("document.edit", document, artifacts),
    )
    assert edited.succeeded and document.read_text() == "supported claim"


def test_document_edit_rejects_parent_swap_after_authorization(
        artifacts, tmp_path, monkeypatch):
    inside = artifacts / "inside"
    inside.mkdir()
    target = inside / "note.md"
    target.write_text("inside")
    outside = tmp_path / "outside"
    outside.mkdir()
    outside_target = outside / "note.md"
    outside_target.write_text("outside")
    tools = document_tools.DocumentTools((artifacts,))
    approval = approve("document.edit", target, artifacts)
    authorize = tools.artifacts.runtime.authorize

    def swap_after_authorization(*args, **kwargs):
        actions = authorize(*args, **kwargs)
        inside.rename(artifacts / "displaced")
        inside.symlink_to(outside, target_is_directory=True)
        return actions

    monkeypatch.setattr(tools.artifacts.runtime, "authorize", swap_after_authorization)
    with pytest.raises(OSError):
        tools.edit(target, [("inside", "changed")], approval_id=approval)
    assert outside_target.read_text() == "outside"


def test_document_edit_rejects_file_identity_swap_before_commit(
        artifacts, monkeypatch):
    target = artifacts / "note.md"
    target.write_text("original")
    tools = document_tools.DocumentTools((artifacts,))
    approval = approve("document.edit", target, artifacts)
    writer = tools.artifacts.writer
    swapped = False

    def swap_before_writer(*args, **kwargs):
        nonlocal swapped
        if not swapped:
            swapped = True
            target.rename(artifacts / "displaced.md")
            target.write_text("replacement")
        return writer(*args, **kwargs)

    monkeypatch.setattr(tools.artifacts, "writer", swap_before_writer)
    with pytest.raises(RuntimeError, match="object changed"):
        tools.edit(target, [("original", "updated")], approval_id=approval)
    assert target.read_text() == "replacement"
    assert (artifacts / "displaced.md").read_text() == "original"


def test_pdf_capabilities_are_truthful_and_info_is_structured(monkeypatch, artifacts):
    target = artifacts / "fixture.pdf"
    target.write_bytes(b"%PDF fixture")
    monkeypatch.setattr(artifact_tools.shutil, "which", lambda name: f"/usr/bin/{name}")
    runner = lambda argv, **kwargs: subprocess.CompletedProcess(
        argv, 0, "Pages: 2\nPDF version: 1.7\n", ""
    )
    tools = artifact_tools.PDFTools((artifacts,), runner=runner)
    result = tools.info(target)
    assert result.succeeded and result.data["metadata"]["Pages"] == "2"
    assert tools.capabilities()["redaction"] is False


def test_image_unavailable_is_reported_without_fabrication(monkeypatch, artifacts):
    target = artifacts / "image.png"
    target.write_bytes(b"not needed")
    monkeypatch.setattr(media_tools.ImageTools, "_image",
                        staticmethod(lambda: (_ for _ in ()).throw(
                            RuntimeError("Pillow image backend is unavailable"))))
    result = media_tools.ImageTools((artifacts,)).info(target)
    assert result.state == "unavailable" and not result.succeeded


def test_optional_pdf_spreadsheet_and_image_backends_real_round_trip(artifacts):
    pypdf = pytest.importorskip("pypdf")
    pytest.importorskip("openpyxl")
    Image = pytest.importorskip("PIL.Image")

    source_pdf = artifacts / "source.pdf"
    writer = pypdf.PdfWriter()
    writer.add_blank_page(width=100, height=100)
    writer.add_blank_page(width=100, height=100)
    with source_pdf.open("wb") as handle:
        writer.write(handle)
    output_pdf = artifacts / "reordered.pdf"
    pdf = artifact_tools.PDFTools((artifacts,))
    transformed = pdf.reorder(source_pdf, output_pdf, [1, 0], approval_ids={
        "write-0": approve("pdf.reorder", output_pdf, artifacts),
    })
    assert transformed.succeeded and transformed.data["pages"] == 2
    parts = artifacts / "parts"
    split = pdf.split(source_pdf, parts, ([0], [1]), approval_ids={
        "write-0": approve("pdf.split", parts, artifacts),
    })
    assert split.succeeded and [part["pages"] for part in split.data["parts"]] == [1, 1]

    from openpyxl import Workbook, load_workbook
    workbook_path = artifacts / "book.xlsx"
    workbook = Workbook()
    workbook.save(workbook_path)
    workbook.close()
    sheets = document_tools.SpreadsheetTools((artifacts,))
    added = sheets.add_sheet(workbook_path, "Data",
                             approval_id=approve("spreadsheet.add_sheet", workbook_path, artifacts))
    assert added.succeeded and "Data" in load_workbook(workbook_path, read_only=True).sheetnames

    source_image = artifacts / "source.png"
    Image.new("RGB", (20, 10), "blue").save(source_image)
    output_image = artifacts / "small.jpg"
    resized = media_tools.ImageTools((artifacts,)).transform(
        "resize", source_image, output_image, size=(5, 4), format="JPEG",
        approval_ids={"write-0": approve("image.resize", output_image, artifacts)},
    )
    assert resized.succeeded and resized.data["dimensions"] == [5, 4]
