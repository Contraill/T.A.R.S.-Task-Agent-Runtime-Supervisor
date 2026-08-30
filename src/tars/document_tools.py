from __future__ import annotations

import csv
import hashlib
import importlib.util
from pathlib import Path
import shutil
import subprocess

from .artifact_tools import ArtifactRuntime
from .tool_core import ToolResult


class DocumentTools:
    TEXT_FORMATS = {".txt", ".md", ".rst", ".csv", ".json", ".html", ".htm"}

    def __init__(self, roots, *, runtime=None, runner=subprocess.run):
        self.artifacts = ArtifactRuntime(roots, runtime=runtime)
        self.runner = runner

    def capabilities(self):
        libreoffice = bool(shutil.which("libreoffice"))
        return {"inspect": sorted(self.TEXT_FORMATS | ({".docx"} if importlib.util.find_spec("docx") else set())),
                "extract": sorted(self.TEXT_FORMATS | ({".docx"} if importlib.util.find_spec("docx") else set())),
                "convert": libreoffice, "edit": sorted(self.TEXT_FORMATS),
                "universal_support": False}

    def inspect(self, path, *, task_id=None, session_id=None):
        actions, reads, _ = self.artifacts.authorize("document.inspect", (path,),
                                                     task_id=task_id, session_id=session_id)
        target = reads[0]
        suffix = target.suffix.casefold()
        supported = suffix in self.capabilities()["inspect"]
        data = {"path": str(target), "format": suffix.removeprefix("."),
                "bytes": target.stat().st_size, "supported": supported,
                "sha256": hashlib.sha256(target.read_bytes()).hexdigest()}
        return self.artifacts.result("document.inspect", actions, data, task_id=task_id,
                                     evidence_source=target,
                                     state="succeeded" if supported else "unavailable",
                                     error="" if supported else "document format is unsupported")

    def extract(self, path, *, task_id=None, session_id=None):
        actions, reads, _ = self.artifacts.authorize("document.extract", (path,),
                                                     task_id=task_id, session_id=session_id)
        target = reads[0]
        try:
            if target.suffix.casefold() in self.TEXT_FORMATS:
                text = target.read_text(errors="replace")
            elif target.suffix.casefold() == ".docx" and importlib.util.find_spec("docx"):
                from docx import Document
                document = Document(str(target))
                text = "\n".join(paragraph.text for paragraph in document.paragraphs)
            else:
                data = {"path": str(target), "available": False,
                        "reason": "document extraction backend unavailable"}
                return self.artifacts.result("document.extract", actions, data, task_id=task_id,
                                             evidence_source=target, state="unavailable",
                                             error=data["reason"])
            data = {"path": str(target), "text": text, "characters": len(text)}
        except Exception as exc:
            self.artifacts.runtime.finish(actions, state="failed", result={"error": str(exc)})
            raise
        return self.artifacts.result("document.extract", actions, data, task_id=task_id,
                                     evidence_source=target)

    def edit(self, path, replacements, *, approval_id=None, task_id=None, session_id=None):
        actions, reads, writes = self.artifacts.authorize(
            "document.edit", (path,), (path,), approval_ids={"write-0": approval_id},
            task_id=task_id, session_id=session_id,
            arguments={"replacements": len(replacements)},
        )
        target = writes[0]
        if target.suffix.casefold() not in self.TEXT_FORMATS:
            self.artifacts.runtime.finish(actions, state="failed",
                                          result={"error": "format is not safely editable"})
            return ToolResult("document.edit", "unavailable", {"path": str(target)},
                              "format is not safely editable",
                              action_ids=tuple(action.id for action in actions))
        original = target.read_text()
        updated = original
        try:
            for old, new in replacements:
                count = updated.count(old)
                if count != 1:
                    raise ValueError(f"edit context must match exactly once; matched {count}")
                updated = updated.replace(old, new, 1)
            target.write_text(updated)
            data = {"path": str(target), "replacements": len(replacements),
                    "before_sha256": hashlib.sha256(original.encode()).hexdigest(),
                    "after_sha256": hashlib.sha256(updated.encode()).hexdigest()}
        except Exception as exc:
            self.artifacts.runtime.finish(actions, state="failed", result={"error": str(exc)})
            raise
        return self.artifacts.result("document.edit", actions, data, task_id=task_id,
                                     evidence_source=target)

    def convert(self, path, output_dir, output_format, *, approval_ids=None,
                task_id=None, session_id=None):
        actions, reads, writes = self.artifacts.authorize(
            "document.convert", (path,), (output_dir,), approval_ids=approval_ids,
            task_id=task_id, session_id=session_id, arguments={"format": output_format},
        )
        binary = shutil.which("libreoffice")
        if not binary:
            data = {"available": False, "reason": "LibreOffice conversion backend unavailable"}
            return self.artifacts.result("document.convert", actions, data,
                                         state="unavailable", error=data["reason"])
        destination = writes[0]
        destination.mkdir(parents=True, exist_ok=True)
        proc = self.runner([binary, "--headless", "--convert-to", output_format,
                            "--outdir", str(destination), str(reads[0])],
                           capture_output=True, text=True, check=False)
        primary_format = output_format.split(":", 1)[0]
        expected = destination / f"{reads[0].stem}.{primary_format}"
        outputs = [str(expected)] if expected.is_file() else []
        data = {"source": str(reads[0]), "output_dir": str(destination), "files": outputs,
                "exit_code": proc.returncode, "stdout": proc.stdout, "stderr": proc.stderr,
                "regenerated": True}
        verified = proc.returncode == 0 and bool(outputs)
        return self.artifacts.result("document.convert", actions, data, task_id=task_id,
                                     evidence_source=destination,
                                     state="succeeded" if verified else "failed",
                                     error="" if verified else proc.stderr or "no output generated")


class SpreadsheetTools:
    def __init__(self, roots, *, runtime=None):
        self.artifacts = ArtifactRuntime(roots, runtime=runtime)

    def capabilities(self):
        advanced = bool(importlib.util.find_spec("openpyxl"))
        return {"csv": {"inspect": True, "read_range": True, "write_range": True,
                         "add_sheet": False, "formulas": False, "export": True},
                "xlsx": {"inspect": advanced, "read_range": advanced,
                          "write_range": advanced, "add_sheet": advanced,
                          "formulas": advanced, "export": advanced}}

    @staticmethod
    def _a1(cell):
        import re
        match = re.fullmatch(r"([A-Za-z]+)([1-9][0-9]*)", cell)
        if not match:
            raise ValueError(f"invalid cell reference: {cell}")
        column = 0
        for char in match.group(1).upper():
            column = column * 26 + ord(char) - 64
        return int(match.group(2)) - 1, column - 1

    @classmethod
    def _range(cls, value):
        start, _, end = value.partition(":")
        row1, col1 = cls._a1(start)
        row2, col2 = cls._a1(end or start)
        if row2 < row1 or col2 < col1:
            raise ValueError("spreadsheet range is reversed")
        return row1, col1, row2, col2

    def inspect(self, path, *, task_id=None, session_id=None):
        actions, reads, _ = self.artifacts.authorize("spreadsheet.inspect", (path,),
                                                     task_id=task_id, session_id=session_id)
        target = reads[0]
        if target.suffix.casefold() == ".csv":
            with target.open(newline="") as handle:
                rows = list(csv.reader(handle))
            data = {"path": str(target), "format": "csv", "sheets": ["Sheet1"],
                    "rows": len(rows), "columns": max((len(row) for row in rows), default=0)}
        elif target.suffix.casefold() == ".xlsx" and importlib.util.find_spec("openpyxl"):
            from openpyxl import load_workbook
            workbook = load_workbook(target, read_only=True, data_only=False)
            data = {"path": str(target), "format": "xlsx", "sheets": workbook.sheetnames,
                    "dimensions": {sheet.title: sheet.calculate_dimension()
                                   for sheet in workbook.worksheets}}
            workbook.close()
        else:
            data = {"path": str(target), "available": False,
                    "reason": "spreadsheet format/backend unavailable"}
            return self.artifacts.result("spreadsheet.inspect", actions, data,
                                         state="unavailable", error=data["reason"])
        return self.artifacts.result("spreadsheet.inspect", actions, data, task_id=task_id,
                                     evidence_source=target)

    def read_range(self, path, cell_range, *, sheet=None, formulas=True,
                   task_id=None, session_id=None):
        actions, reads, _ = self.artifacts.authorize("spreadsheet.read_range", (path,),
                                                     task_id=task_id, session_id=session_id)
        target = reads[0]
        row1, col1, row2, col2 = self._range(cell_range)
        if target.suffix.casefold() == ".csv":
            with target.open(newline="") as handle:
                rows = list(csv.reader(handle))
            values = [[rows[r][c] if r < len(rows) and c < len(rows[r]) else ""
                       for c in range(col1, col2 + 1)] for r in range(row1, row2 + 1)]
        elif target.suffix.casefold() == ".xlsx" and importlib.util.find_spec("openpyxl"):
            from openpyxl import load_workbook
            workbook = load_workbook(target, read_only=True, data_only=not formulas)
            worksheet = workbook[sheet] if sheet else workbook.active
            values = [[worksheet.cell(r + 1, c + 1).value for c in range(col1, col2 + 1)]
                      for r in range(row1, row2 + 1)]
            workbook.close()
        else:
            data = {"available": False, "reason": "spreadsheet backend unavailable"}
            return self.artifacts.result("spreadsheet.read_range", actions, data,
                                         state="unavailable", error=data["reason"])
        data = {"path": str(target), "sheet": sheet or "Sheet1", "range": cell_range,
                "values": values, "formulas": formulas}
        return self.artifacts.result("spreadsheet.read_range", actions, data, task_id=task_id,
                                     evidence_source=target)

    def write_range(self, path, cell_range, values, *, sheet=None, approval_id=None,
                    task_id=None, session_id=None):
        actions, reads, writes = self.artifacts.authorize(
            "spreadsheet.write_range", (path,), (path,), approval_ids={"write-0": approval_id},
            task_id=task_id, session_id=session_id, arguments={"range": cell_range},
        )
        target = writes[0]
        row1, col1, row2, col2 = self._range(cell_range)
        if len(values) != row2 - row1 + 1 or any(len(row) != col2 - col1 + 1 for row in values):
            self.artifacts.runtime.finish(actions, state="failed",
                                          result={"error": "values do not match range"})
            raise ValueError("values do not match spreadsheet range")
        if target.suffix.casefold() == ".csv":
            with target.open(newline="") as handle:
                rows = list(csv.reader(handle))
            while len(rows) <= row2:
                rows.append([])
            for r, values_row in enumerate(values, row1):
                while len(rows[r]) <= col2:
                    rows[r].append("")
                rows[r][col1:col2 + 1] = [str(value) for value in values_row]
            with target.open("w", newline="") as handle:
                csv.writer(handle).writerows(rows)
        elif target.suffix.casefold() == ".xlsx" and importlib.util.find_spec("openpyxl"):
            from openpyxl import load_workbook
            workbook = load_workbook(target)
            worksheet = workbook[sheet] if sheet else workbook.active
            for r, values_row in enumerate(values, row1 + 1):
                for c, value in enumerate(values_row, col1 + 1):
                    worksheet.cell(r, c, value)
            workbook.save(target)
            workbook.close()
        else:
            self.artifacts.runtime.finish(actions, state="failed",
                                          result={"error": "spreadsheet backend unavailable"})
            return ToolResult("spreadsheet.write_range", "unavailable", {},
                              "spreadsheet backend unavailable",
                              action_ids=tuple(action.id for action in actions))
        data = {"path": str(target), "sheet": sheet or "Sheet1", "range": cell_range,
                "sha256": hashlib.sha256(target.read_bytes()).hexdigest()}
        return self.artifacts.result("spreadsheet.write_range", actions, data, task_id=task_id,
                                     evidence_source=target)

    def add_sheet(self, path, name, *, approval_id=None, task_id=None, session_id=None):
        actions, _, writes = self.artifacts.authorize(
            "spreadsheet.add_sheet", (path,), (path,), approval_ids={"write-0": approval_id},
            task_id=task_id, session_id=session_id, arguments={"name": name},
        )
        target = writes[0]
        if target.suffix.casefold() != ".xlsx" or not importlib.util.find_spec("openpyxl"):
            self.artifacts.runtime.finish(actions, state="failed",
                                          result={"error": "xlsx backend unavailable"})
            return ToolResult("spreadsheet.add_sheet", "unavailable", {},
                              "xlsx backend unavailable", action_ids=tuple(a.id for a in actions))
        from openpyxl import load_workbook
        workbook = load_workbook(target)
        workbook.create_sheet(name)
        workbook.save(target)
        workbook.close()
        data = {"path": str(target), "sheet": name, "created": True}
        return self.artifacts.result("spreadsheet.add_sheet", actions, data, task_id=task_id,
                                     evidence_source=target)

    def export(self, path, output, *, sheet=None, approval_ids=None,
               task_id=None, session_id=None):
        actions, reads, writes = self.artifacts.authorize(
            "spreadsheet.export", (path,), (output,), approval_ids=approval_ids,
            task_id=task_id, session_id=session_id,
        )
        source, destination = reads[0], writes[0]
        if destination.suffix.casefold() != ".csv":
            self.artifacts.runtime.finish(actions, state="failed",
                                          result={"error": "only CSV export is supported"})
            return ToolResult("spreadsheet.export", "unavailable", {},
                              "only CSV export is supported", action_ids=tuple(a.id for a in actions))
        if source.suffix.casefold() == ".csv":
            shutil.copy2(source, destination)
        elif source.suffix.casefold() == ".xlsx" and importlib.util.find_spec("openpyxl"):
            from openpyxl import load_workbook
            workbook = load_workbook(source, read_only=True, data_only=False)
            worksheet = workbook[sheet] if sheet else workbook.active
            with destination.open("w", newline="") as handle:
                csv.writer(handle).writerows(row for row in worksheet.iter_rows(values_only=True))
            workbook.close()
        else:
            self.artifacts.runtime.finish(actions, state="failed",
                                          result={"error": "spreadsheet backend unavailable"})
            return ToolResult("spreadsheet.export", "unavailable", {},
                              "spreadsheet backend unavailable", action_ids=tuple(a.id for a in actions))
        data = {"source": str(source), "output": str(destination), "format": "csv",
                "sha256": hashlib.sha256(destination.read_bytes()).hexdigest()}
        return self.artifacts.result("spreadsheet.export", actions, data, task_id=task_id,
                                     evidence_source=destination)

    def formulas(self, path, formulas, *, sheet=None, approval_id=None,
                 task_id=None, session_id=None):
        """Write an A1-to-formula mapping without evaluating formulas in T.A.R.S."""
        if not formulas:
            raise ValueError("at least one formula is required")
        if Path(path).suffix.casefold() != ".xlsx" or not importlib.util.find_spec("openpyxl"):
            return ToolResult("spreadsheet.formulas", "unavailable", {},
                              "xlsx formula backend unavailable")
        actions, _, writes = self.artifacts.authorize(
            "spreadsheet.formulas", (path,), (path,), approval_ids={"write-0": approval_id},
            task_id=task_id, session_id=session_id,
            arguments={"cells": sorted(formulas)},
        )
        from openpyxl import load_workbook
        target = writes[0]
        workbook = load_workbook(target)
        worksheet = workbook[sheet] if sheet else workbook.active
        for cell, formula in formulas.items():
            if not isinstance(formula, str) or not formula.startswith("="):
                self.artifacts.runtime.finish(actions, state="failed",
                                              result={"error": "formula must start with ="})
                raise ValueError("formula must start with =")
            worksheet[cell] = formula
        workbook.save(target)
        workbook.close()
        data = {"path": str(target), "sheet": sheet or "Sheet1",
                "cells": sorted(formulas), "calculated": False,
                "sha256": hashlib.sha256(target.read_bytes()).hexdigest()}
        return self.artifacts.result("spreadsheet.formulas", actions, data, task_id=task_id,
                                     evidence_source=target)
